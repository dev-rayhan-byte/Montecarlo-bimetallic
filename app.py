import traceback
import streamlit as st
import pandas as pd
import numpy as np
import threading
import time
import os
import math
import tempfile
import ast
import base64
from openpyxl import Workbook
from ase.cluster import FaceCenteredCubic, BodyCenteredCubic, HexagonalClosedPacked
from ase.io import write
from ase.neighborlist import build_neighbor_list
import matplotlib.pyplot as plt

# ---- Streamlit Page Config ----
st.set_page_config(page_title="Monte Carlo Nanoparticle Simulator", layout="wide")

# ---- Constants ----
BOLTZMANN_K = 8.617333262e-5
BULK_COORD = 12
lattice_map = {
    'fcc': FaceCenteredCubic,
    'bcc': BodyCenteredCubic,
    'hcp': HexagonalClosedPacked
}

# ---- Helper Functions ----
def symbol_type(sym, A):
    return 'A' if sym == A else 'B'

def calculate_energy(p, A, coeffs):
    nl = build_neighbor_list(p, bothways=True, self_interaction=False)
    count = {k: 0 for k in coeffs}
    for i in range(len(p)):
        t_i = symbol_type(p[i].symbol, A)
        neighbors = nl.get_neighbors(i)[0]
        if len(neighbors) < BULK_COORD:
            count[f'x{t_i}-S'] += 1
        for j in neighbors:
            if i < j:
                t_j = symbol_type(p[j].symbol, A)
                if t_i == t_j:
                    count[f'x{t_i}-{t_j}'] += 1
                else:
                    count['xA-B'] += 1
    return sum(coeffs[k] * count.get(k, 0) for k in count)

def count_surface(p, A):
    nl = build_neighbor_list(p, bothways=True, self_interaction=False)
    total_A, surf, surf_A = 0, 0, 0
    for i in range(len(p)):
        neighbors = nl.get_neighbors(i)[0]
        t = symbol_type(p[i].symbol, A)
        if t == 'A':
            total_A += 1
        if len(neighbors) < BULK_COORD:
            surf += 1
            if t == 'A':
                surf_A += 1
    ratio = surf_A / surf if surf else 0
    return total_A, surf, surf_A, ratio

def run_simulation(params, progress_callback=None):
    A = params['element_A']
    B = params['element_B']
    composition_A = params['composition_A']
    T = params['temperature']
    N_STEPS = params['n_steps']
    SAVE_INTERVAL = params['save_interval']
    LAYERS = params['layers']
    SURFACES = params['surfaces']
    coeffs = params['coefficients']
    lattice_type = params['lattice_type']

    ClusterBuilder = lattice_map.get(lattice_type)
    if ClusterBuilder is None:
        raise ValueError(f"Unsupported lattice type '{lattice_type}'. Choose from 'fcc', 'bcc', or 'hcp'.")

    # Build Initial Particle
    particle = ClusterBuilder(A, surfaces=SURFACES, layers=LAYERS)
    n_atoms = len(particle)
    n_A = int(n_atoms * composition_A)
    indices_A = np.random.choice(range(n_atoms), size=n_A, replace=False)
    for i in range(n_atoms):
        particle[i].symbol = A if i in indices_A else B

    initial_surface_data = count_surface(particle, A)

    # Prepare Storage
    os.makedirs("trajectory", exist_ok=True)
    log = []
    energy = calculate_energy(particle, A, coeffs)
    start_time = time.time()

    # Monte Carlo Loop
    for step in range(1, N_STEPS + 1):
        i = np.random.randint(0, n_atoms)
        neighbors = build_neighbor_list(particle).get_neighbors(i)[0]
        if len(neighbors) == 0:
            continue
        j = np.random.choice(neighbors)
        if particle[i].symbol == particle[j].symbol:
            continue

        trial = particle.copy()
        trial[i].symbol, trial[j].symbol = trial[j].symbol, trial[i].symbol
        dE = calculate_energy(trial, A, coeffs) - energy

        if dE < 0 or np.random.random() < math.exp(-dE / (BOLTZMANN_K * T)):
            particle = trial
            energy += dE

        if step % SAVE_INTERVAL == 0:
            total_A, surf, surf_A, ratio = count_surface(particle, A)
            log.append({
                'Step': step,
                'Energy (eV)': energy,
                f'Total {A}': total_A,
                'Surface Atoms': surf,
                f'{A} on Surface': surf_A,
                f'Surface {A} Ratio': ratio
            })
            if progress_callback:
                progress_callback(step, energy, ratio)

        if step % params.get('snapshot_interval', 500) == 0:
            write(f"trajectory/step_{step:05d}.xyz", particle)

    duration = time.time() - start_time

    # Save Initial and Final XYZ files
    initial_xyz = tempfile.NamedTemporaryFile(delete=False, suffix='.xyz').name
    final_xyz = tempfile.NamedTemporaryFile(delete=False, suffix='.xyz').name
    write(initial_xyz, ClusterBuilder(A, surfaces=SURFACES, layers=LAYERS))
    write(final_xyz, particle)

    # Save Excel Log
    xlsx_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name
    wb = Workbook()
    ws = wb.active
    ws.title = "Simulation Log"

    meta = [
        ("Element A", A), ("Element B", B), ("Composition A", composition_A),
        ("Temperature (K)", T), ("MC Steps", N_STEPS),
        ("Save Interval", SAVE_INTERVAL), ("Total Atoms", n_atoms)
    ]
    for i, (k, v) in enumerate(meta, 1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    header_row = len(meta) + 2
    if log:
        headers = list(log[0].keys())
        for j, h in enumerate(headers, 1):
            ws.cell(row=header_row, column=j, value=h)
        for i, entry in enumerate(log, header_row + 1):
            for j, h in enumerate(headers, 1):
                ws.cell(row=i, column=j, value=entry[h])

    wb.save(xlsx_file)

    return {
        "initial_xyz": initial_xyz,
        "final_xyz": final_xyz,
        "log": pd.DataFrame(log),
        "xlsx_file": xlsx_file,
        "duration": duration,
        "initial_surface_data": initial_surface_data,
        "final_surface_data": count_surface(particle, A)
    }

def make_download_link(path, label=None):
    label = label or os.path.basename(path)
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{os.path.basename(path)}">📥 {label}</a>'
    st.markdown(href, unsafe_allow_html=True)

# ---- Streamlit Sidebar UI ----
st.sidebar.header("Simulation Parameters")

with st.sidebar.expander("Elements & Composition", expanded=True):
    element_A = st.text_input("Element A (e.g., Pt)", value="Pt")
    element_B = st.text_input("Element B (e.g., Ru)", value="Ru")
    composition_A = st.slider("Composition of A (fraction)", 0.0, 1.0, 0.5, step=0.05)

with st.sidebar.expander("Thermodynamics"):
    temperature = st.number_input("Temperature (K)", min_value=1.0, value=250.0, step=1.0)
    n_steps = st.number_input("Monte Carlo Steps", min_value=100, value=2000, step=100)
    save_interval = st.number_input("Log Save Interval", min_value=1, value=200, step=1)
    snapshot_interval = st.number_input("Snapshot Interval", min_value=100, value=500, step=100)

with st.sidebar.expander("Lattice / Geometry"):
    lattice_type = st.selectbox("Lattice Type", ["fcc", "bcc", "hcp"])
    layers = st.multiselect("Layers (x,y,z)", options=[1,2,3,4,5,6,7,8], default=[7,7,7])
    if len(layers) != 3:
        st.warning("Please select exactly 3 values for layers.")
        layers = [7,7,7]
    surfaces_input = st.text_input("Surfaces (list of tuples)", value="[(1,1,1),(1,1,1),(1,1,0)]")
    try:
        surfaces_parsed = ast.literal_eval(surfaces_input)
    except:
        st.error("Invalid surface input! Using default [(1,1,1),(1,1,1),(1,1,0)]")
        surfaces_parsed = [(1,1,1),(1,1,1),(1,1,0)]

with st.sidebar.expander("Energy Coefficients"):
    coeffs = {
        'xA-A': st.number_input("xA-A", value=-0.022078),
        'xB-B': st.number_input("xB-B", value=-0.150000),
        'xA-B': st.number_input("xA-B", value=-0.109575),
        'xA-S': st.number_input("xA-S", value=-0.250717),
        'xB-S': st.number_input("xB-S", value=-0.300000),
        'xA-A-out': st.number_input("xA-A-out", value=0.184150),
        'xB-B-out': st.number_input("xB-B-out", value=0.332228),
        'xA-B-out': st.number_input("xA-B-out", value=0.051042),
    }

run_button = st.sidebar.button("▶️ Run Simulation")
progress_bar = st.sidebar.progress(0)
status_placeholder = st.empty()

# ---- Simulation Execution ----
if run_button:
    params = {
        'element_A': element_A,
        'element_B': element_B,
        'composition_A': composition_A,
        'temperature': temperature,
        'n_steps': int(n_steps),
        'save_interval': int(save_interval),
        'snapshot_interval': int(snapshot_interval),
        'layers': tuple(layers),
        'surfaces': surfaces_parsed,
        'coefficients': coeffs,
        'lattice_type': lattice_type
    }

    progress_state = {"step": 0, "energy": None, "ratio": None, "done": False, "error": None}
    result_holder = {}

    def progress_cb(step, energy, ratio):
        progress_state["step"] = step
        progress_state["energy"] = energy
        progress_state["ratio"] = ratio

    def simulation_thread():
        try:
            result = run_simulation(params, progress_callback=progress_cb)
            result_holder.update(result)
        except Exception:
            result_holder["error"] = traceback.format_exc()
        progress_state["done"] = True

    thread = threading.Thread(target=simulation_thread)
    thread.start()

    with st.spinner("🌀 Running Monte Carlo simulation... please wait..."):
        while not progress_state["done"]:
            if progress_state["step"] > 0:
                pct = min(progress_state["step"] / params['n_steps'], 1.0)
                progress_bar.progress(pct)
                status_placeholder.markdown(
                    f"**Step:** {progress_state['step']} | "
                    f"**Energy:** {progress_state['energy']:.4f} eV | "
                    f"**Surface {element_A} Ratio:** {progress_state['ratio']:.4f}"
                )
            time.sleep(0.5)

    # ---- Results Display ----
    if "error" in result_holder:
        st.error("❌ Simulation failed.")
        with st.expander("🔍 Error Details"):
            st.code(result_holder["error"])
    else:
        st.success("✅ Simulation completed.")
        res = result_holder
        df_log = res["log"]

        st.subheader("Simulation Summary")
        col1, col2, col3 = st.columns(3)
        col1.metric("Duration (s)", f"{res['duration']:.1f}")
        init_total, init_surf, init_surf_A, init_ratio = res["initial_surface_data"]
        final_total, final_surf, final_surf_A, final_ratio = res["final_surface_data"]
        col2.metric(f"Initial Surface {element_A} Ratio", f"{init_ratio:.4f}")
        col3.metric(f"Final Surface {element_A} Ratio", f"{final_ratio:.4f}")

        st.subheader("Evolution Plots")
        fig1, ax1 = plt.subplots()
        ax1.plot(df_log["Step"], df_log["Energy (eV)"])
        ax1.set_xlabel("MC Step")
        ax1.set_ylabel("Energy (eV)")
        ax1.grid(True)
        st.pyplot(fig1)

        fig2, ax2 = plt.subplots()
        ax2.plot(df_log["Step"], df_log[f"Surface {element_A} Ratio"], color="orange")
        ax2.set_xlabel("MC Step")
        ax2.set_ylabel(f"Surface {element_A} Ratio")
        ax2.grid(True)
        st.pyplot(fig2)

        st.subheader("Download Artifacts")
        with st.expander("Files"):
            make_download_link(res["initial_xyz"], "Initial structure (.xyz)")
            make_download_link(res["final_xyz"], "Final structure (.xyz)")
            make_download_link(res["xlsx_file"], "Simulation log (.xlsx)")

        st.subheader("Raw Log Data")
        st.dataframe(df_log)
